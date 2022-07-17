import codecs
from email import message
from hashlib import new
import os
from tkinter.messagebox import NO
from turtle import home
from unicodedata import name
from xml.dom import minidom
from django.http import Http404
from django.http import FileResponse
import openpyxl
from pypinyin import pinyin, lazy_pinyin, Style
from django.db import connection

from rest_framework import status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.decorators import api_view
from rest_framework.viewsets import ViewSet
from rest_framework.viewsets import ModelViewSet
from django.conf import settings
import xml.etree.ElementTree as ET
import chardet

from .models import AppliedDgStd, AppliedGStds, AppliedSpStd, DiagStd, GStd, Specialty1, Specialty2, SpecialtyStd, UploadModel#, Specialty3
from .serializers import DiagSerializer, DiagStdSerializer, GStdSerializer, GeneralSerializer, SpecialtyStdSerializer, UploadSerializer, Specialty1Serializer, Specialty2Serializer, Specialty3Serializer

from django.db.models import Q # database查询

import xlrd
import openpyxl
import json
import xmltodict

class UploadXlsViewSet(ModelViewSet):
    queryset = UploadModel.objects.all()
    def create(self, request):
        file_dir = None
        print('UploadViewSet.create()')
        upload_serializer = UploadSerializer(data=request.data)
        print('validation:', upload_serializer.is_valid())
        if(upload_serializer.is_valid()):
            upload = upload_serializer.save()
        file_name = str(upload.file).split('/')[1]
        file_dir = str(settings.MEDIA_ROOT)+'\\uploads\\'+file_name
        print('file_dir:', file_dir)

        output = dict(name = file_name, specialty1 = [])

        file_type = file_name.split('.')[1]
        if(file_type=='xls'):
            # xlrd.Book.encoding = "gbk"
            file = xlrd.open_workbook(file_dir)
            sheet = file.sheets()[0]
            print(sheet)
            print(sheet.nrows)
            for i in range(1,sheet.nrows):
                val = str(sheet.cell_value(i,0)).strip(' ')
                label = sheet.cell_value(i,1).strip(' ')
                print(val+' -- '+label+'     >> '+str(len(val)))
                if len(val)==2:
                    temp = dict(value = val, label=label)
                    output['specialty1'].append(temp)
                elif len(val)==4:
                    father_val = val[0:2]
                    temp_val = val[2:4]
                    temp = dict(value = temp_val, label = label)
                    for sp1 in output['specialty1']:
                        if sp1['value'] == father_val:
                            if 'specialty2' not in sp1:
                                sp1['specialty2'] = []
                            sp1['specialty2'].append(temp)
                elif len(val)==6:
                    grandpa_val = val[0:2]
                    father_val = val[2:4]
                    temp_val = val[4:6]
                    temp = dict(value = temp_val, label = label)
                    for sp1 in output['specialty1']:
                        if sp1['value'] == grandpa_val:
                            for sp2 in sp1['specialty2']:
                                if sp2['value'] == father_val:
                                    if 'specialty3' not in sp2:
                                        sp2['specialty3'] = []
                                    sp2['specialty3'].append(temp)
        elif(file_type=='xlsx'):
            file = openpyxl.load_workbook(file_dir)
            sheets = file.sheetnames
            sheet = file[sheets[0]]
            print('max_row:',sheet.max_row)
            for i in range(2,sheet.max_row + 1):
                if sheet.cell(column=1, row=i).value==None or sheet.cell(column=2, row=i).value==None:
                    continue
                val = str(sheet.cell(column=1, row=i).value).strip(' ')
                label = sheet.cell(column=2, row=i).value.strip(' ')
                print(val+' -- '+label+'     >> '+str(len(val)))
                if len(val)==2:
                    temp = dict(value = val, label=label)
                    output['specialty1'].append(temp)
                elif len(val)==4:
                    father_val = val[0:2]
                    temp_val = val[2:4]
                    temp = dict(value = temp_val, label = label)
                    for sp1 in output['specialty1']:
                        if sp1['value'] == father_val:
                            if 'specialty2' not in sp1:
                                sp1['specialty2'] = []
                            sp1['specialty2'].append(temp)
                elif len(val)==6:
                    grandpa_val = val[0:2]
                    father_val = val[2:4]
                    temp_val = val[4:6]
                    temp = dict(value = temp_val, label = label)
                    for sp1 in output['specialty1']:
                        if sp1['value'] == grandpa_val:
                            for sp2 in sp1['specialty2']:
                                if sp2['value'] == father_val:
                                    if 'specialty3' not in sp2:
                                        sp2['specialty3'] = []
                                    sp2['specialty3'].append(temp)
        else:
            return Response("有问题啊")

        # print(output)
        specialtystd_serializer = SpecialtyStdSerializer(data=output)
        print('validation:', specialtystd_serializer.is_valid())
        if specialtystd_serializer.is_valid():
            specialtystd_serializer.save()
        else:
            print(specialtystd_serializer.errors)
        print('giao')
        response = "POST API: 你一上传了一个文件:{}".format(file_name)
        return Response(response)

class UploadJsonViewSet(ModelViewSet):
    queryset = UploadModel.objects.all()
    def create(self, request):
        file_dir = None # 文件存储路径
        json_str = None # 读取的json
        print('UploadJsonViewSet.create()')
        upload_serializer = UploadSerializer(data=request.data)
        print('validation:', upload_serializer.is_valid())
        if(upload_serializer.is_valid()):
            upload = upload_serializer.save()
        file_name = str(upload.file).split('/')[1]
        file_dir = str(settings.MEDIA_ROOT)+'\\uploads\\'+file_name
        print('file_dir:', file_dir)
        # 打开json文件
        json_str = None
        with open(file_dir, encoding = 'utf-8') as f:
            json_str = json.load(f)
        if json_str:
            json_str['name'] = file_name
            # print(json_str)
            specialtystd_serializer = SpecialtyStdSerializer(data=json_str)
            print('validation:', specialtystd_serializer.is_valid())
            if specialtystd_serializer.is_valid():
                specialtystd_serializer.save()
            else:
                print(specialtystd_serializer.errors)
            print('giao')
        response = "POST API: 你一上传了一个文件:{}".format(file_name)

        # 砖码并录入

        return Response(response)

class UploadDgXlsViewSet(ModelViewSet):
    queryset = UploadModel.objects.all()
    def create(self, request):
        file_dir = None
        print('UploadDgXlsViewSet.create()')
        upload_serializer = UploadSerializer(data=request.data)
        print('validation:', upload_serializer.is_valid())
        if(upload_serializer.is_valid()):
            upload = upload_serializer.save()
        file_name = str(upload.file).split('/')[1]
        print('file_name:', file_name)
        file_dir = str(settings.MEDIA_ROOT)+'\\uploads\\'+file_name
        print('file_dir:', file_dir)

        output = dict(name = file_name, diag = [])

        file_type = file_name.split('.')[1]
        print('file_type:',file_type)
        if(file_type=='xls'):
            file = xlrd.open_workbook(file_dir)
            sheet = file.sheets()[0]
            print('sheet:',sheet)
            print(sheet.nrows)
            for i in range(1,sheet.nrows):
                code = str(sheet.cell_value(i,0)).strip(' ')
                label = sheet.cell_value(i,1).strip(' ')
                temp = dict(code=code,label=label)
                temp['pinyin']=''.join(map(lambda p:p[0], pinyin(temp['label'], style=Style.NORMAL)))
                temp['pinyin_cap']=''.join(map(lambda p:p[0], pinyin(temp['label'], style=Style.FIRST_LETTER)))
                temp['description']=''
                output['diag'].append(temp)
        elif(file_type=='xlsx'):
            file = openpyxl.load_workbook(file_dir)
            sheets = file.sheetnames
            sheet = file[sheets[0]]
            print('max_row:',sheet.max_row)
            for i in range(2,sheet.max_row + 1):
                if sheet.cell(column=1, row=i).value==None or sheet.cell(column=2, row=i).value==None:
                    continue
                code = str(sheet.cell(column=1, row=i).value).strip(' ')
                label = sheet.cell(column=2, row=i).value.strip(' ')
                print(code+' -- '+label+'     >> '+str(len(code)))
                temp = dict(code=code,label=label,\
                    pinyin=''.join(map(lambda p:p[0],\
                        pinyin(temp['label'], style=Style.NORMAL))),\
                    pinyin_cap=''.join(map(lambda p:p[0],\
                        pinyin(temp['label'], style=Style.FIRST_LETTER))),\
                    description='')
                output['diag'].append(temp)
        else:
            return Response("有问题啊")
        # print(output)
        diagstd_serializer = DiagStdSerializer(data=output)
        print('validation:', diagstd_serializer.is_valid())
        if diagstd_serializer.is_valid():
            diagstd_serializer.save()
        else:
            print(diagstd_serializer.errors)
        print('giao')
        response = "POST API: 你一上传了一个文件:{}".format(file_name)
        return Response(response)

class UploadDgJsonViewSet(ModelViewSet):
    queryset = UploadModel.objects.all()
    def create(self, request):
        file_dir = None # 文件存储路径
        json_str = None # 读取的json
        print('UploadDgJsonViewSet.create()')
        upload_serializer = UploadSerializer(data=request.data)
        print('validation:', upload_serializer.is_valid())
        if(upload_serializer.is_valid()):
            upload = upload_serializer.save()
        file_name = str(upload.file).split('/')[1]
        file_dir = str(settings.MEDIA_ROOT)+'\\uploads\\'+file_name
        print('file_dir:', file_dir)
        # 打开json文件
        json_str = None
        with open(file_dir, encoding = 'utf-8') as f:
            json_str = json.load(f)
        if json_str:
            json_str['name'] = file_name
            for d in json_str['diag']:
                d['pinyin'] = ''.join(map(lambda p:p[0],\
                    pinyin(d['label'], style=Style.NORMAL)))
                d['pinyin_cap'] = ''.join(map(lambda p:p[0],\
                    pinyin(d['label'], style=Style.FIRST_LETTER)))
                d['description'] = ''
            diagstd_serializer = DiagStdSerializer(data=json_str)
            print('validation:', diagstd_serializer.is_valid())
            if diagstd_serializer.is_valid():
                diagstd_serializer.save()
            else:
                print(diagstd_serializer.errors)
            print('giao')
        response = "POST API: 你一上传了一个文件:{}".format(file_name)
        return Response(response)

@api_view(['GET'])
def getJsonExample(request):
    file=open(str(settings.MEDIA_ROOT)+'\\examples\\example.json','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['COntent-Disposition'] = 'attachment; filename="example.json"'
    return response

@api_view(['GET'])
def getXlsExample(request):
    file=open(str(settings.MEDIA_ROOT)+'\\examples\\example.xls','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment; filename="exaple.xls"'
    return response

@api_view(['GET'])
def getXlsxExample(request):
    file=open(str(settings.MEDIA_ROOT)+'\\examples\\example.xlsx','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    return response

@api_view(['GET'])
def getDgJsonExample(request):
    file=open(str(settings.MEDIA_ROOT)+'\\examples\\dg_example.json','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['COntent-Disposition'] = 'attachment; filename="example.json"'
    return response

@api_view(['GET'])
def getDgXlsExample(request):
    file=open(str(settings.MEDIA_ROOT)+'\\examples\\dg_example.xls','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment; filename="exaple.xls"'
    return response

@api_view(['GET'])
def getDgXlsxExample(request):
    file=open(str(settings.MEDIA_ROOT)+'\\examples\\dg_example.xlsx','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment; filename="example.xlsx"'
    return response

@api_view(['GET'])
def getAppliedSpStd(request):
    if not AppliedSpStd.objects.exists():
        print('---- creating AppliedStd')
        if not SpecialtyStd.objects.exists():
            spstd = SpecialtyStd.objects.create(name='默认标准')
        else:
            spstd = SpecialtyStd.objects.filter()[:1].get()
        print('spstd got:',spstd)
        applied_spstd = AppliedSpStd.objects.create(spstd=spstd)
    else:
        applied_spstd = AppliedSpStd.objects.filter()[:1].get()
    print('appliedSpStd got:',applied_spstd)
    response = SpecialtyStdSerializer(applied_spstd.spstd)
    print(type(response.data))
    sp1s = response.data['specialty1']
    id = response.data['id']
    for sp1 in sp1s:
        if 'specialty2' in sp1.keys():
            sp2s = sp1['specialty2']
            for sp2 in sp2s:
                if 'specialty3' in sp2.keys():
                    sp2['children'] = sp2.pop('specialty3')
            sp1['children'] = sp1.pop('specialty2')
    return Response({'id':id,'specialties':sp1s})

@api_view(['POST'])
def setAppliedSpStd(request, format=None):
    print('setAppliedSpStd()')
    print('request.data',request.data)
    id = request.data['id']
    spstd = SpecialtyStd.objects.get(id__exact=id)
    if not AppliedSpStd.objects.exists():
        print('---- creating AppliedStd')
        applied_spstd = AppliedSpStd.objects.create(spstd=spstd)
    else:
        applied_spstd = AppliedSpStd.objects.filter()[:1].get()
        applied_spstd.spstd = spstd
        applied_spstd.save()
    print('appliedSpStd got:',applied_spstd)
    
    response = SpecialtyStdSerializer(applied_spstd.spstd)
    return Response(response.data)

@api_view(['POST'])
def postSpecialtyStd(request, format=None):
    data = request.data['form']
    response = None
    print('data:', data)
    specialtystd_serializer = SpecialtyStdSerializer(data=data)
    print('validation:', specialtystd_serializer.is_valid())
    if specialtystd_serializer.is_valid():
        response = SpecialtyStdSerializer(specialtystd_serializer.save())
    else:
        print(specialtystd_serializer.errors)
    print('giao')
    print(response)
    print(type(response))
    return Response({"spstd":response.data})

@api_view(['POST'])
def postSpecialty1(request, format=None):
    data = request.data['form']
    # {
    #     value: '',
    #     label: '',
    #     description: '',
    #     specialty2
    # }
    print('data:',data)
    specialty1_serializer = Specialty1Serializer(data=data)
    print('validation:',specialty1_serializer.is_valid())
    if specialty1_serializer.is_valid():
        specialty1_serializer.save()
    else:
        print(specialty1_serializer.errors)
    print('giao')
    
    return Response({"request.data":request.data})

@api_view(['POST'])
def postSpecialty2(request, sp1_value, format=None):
    data = request.data['form']
    specialty2_serializer = Specialty2Serializer(data=data, context={"sp1_value":sp1_value})
    print('validation:',specialty2_serializer.is_valid())
    if specialty2_serializer.is_valid():
        specialty2_serializer.save()
    else:
        print(specialty2_serializer.errors)
    print('giao')
    
    return Response({"request.data":request.data})

@api_view(['POST'])
def postSpecialty3(request, sp2_value, format=None):
    data = request.data['form']
    print('going to sp3 serializer')
    specialty3_serializer = Specialty3Serializer(data=data, context={"sp2_value":sp2_value})
    if specialty3_serializer.is_valid():
        specialty3_serializer.save()
    else:
        print(specialty3_serializer.errors)

    return Response({"request.data":request.data})

@api_view(['POST'])
def updateSpecialtyStd(request):
    data = request.data
    print(data)
    print('id:', data['id'])
    print('name:', data['name'])
    print('spstd:',SpecialtyStd.objects.all())
    # instance = SpecialtyStd.objects.get(name=name)
    instance = SpecialtyStd.objects.get(id__exact=data['id'])
    print('instance:',instance)
    # instance.delete()
    specialtystd_serializer = SpecialtyStdSerializer(instance, data)
    print('new spstd valid:',specialtystd_serializer.is_valid())
    if specialtystd_serializer.is_valid():
        specialtystd_serializer.save()
    else:
        print(specialtystd_serializer.errors)
    print('giao')
    
    return Response({"request.data":request.data})

@api_view(['POST'])
def removeSpecialtyStd(request):
    id = request.data['id']
    print('id', id)
    instance = SpecialtyStd.objects.get(id__exact=id)
    if instance:
        print(instance)
        instance.delete()
        print('deleted')
        return Response({"result":id+" removed"})
    else:
        return Response({"result":"something went wrong"})

class ViewSpStd(APIView):
    def get(self, request, format=None):
        id = request.GET.get('id')
        print('spstd w/ id:', id)
        stand = SpecialtyStd.objects.get(id__exact=id)
        print(stand.id)
        serializer = SpecialtyStdSerializer(stand, many=False)

        return Response({'spstd':serializer.data})

class SpStdList(APIView):
    def get(self, request, format=None):
        print('@getSpStdList()')
        spstd_list = SpecialtyStd.objects.all()
        # print(spstd_list)
        serializer = SpecialtyStdSerializer(spstd_list, many=True)
        # print(serializer.data)
        return Response({'spstd_list':serializer.data})

# below are for Diag
@api_view(['GET'])
def getAppliedDgStd(request):
    if not AppliedDgStd.objects.exists():
        print('---- creating AppliedDgStd')
        if not DiagStd.objects.exists():
            dgstd = DiagStd.objects.create(name='默认标准')
        else:
            dgstd = DiagStd.objects.filter()[:1].get()
        print('dgstd got:',dgstd)
        applied_dgstd = AppliedDgStd.objects.create(dgstd=dgstd)
        print('test')
    else:
        applied_dgstd = AppliedDgStd.objects.filter()[:1].get()
    print('applied_stds got:',applied_dgstd)
    response = DiagStdSerializer(applied_dgstd.dgstd)
    print(response.data)
    diags = response.data['diag']
    id = response.data['id']
    return Response({'id':id,'diagnoses':diags})

@api_view(['POST'])
def setAppliedDgStd(request, format=None):
    print('setAppliedDgStd()')
    print('request.data',request.data)
    id = request.data['id']
    dgstd = DiagStd.objects.get(id__exact=id)
    if not AppliedDgStd.objects.exists():
        print('---- creating AppliedStd')
        applied_dgstd = AppliedDgStd.objects.create(dgstd=dgstd)
    else:
        applied_dgstd = AppliedDgStd.objects.filter()[:1].get()
        applied_dgstd.dgstd = dgstd
        applied_dgstd.save()
    print('applied_dgstd got:',applied_dgstd)
    
    response = DiagStdSerializer(applied_dgstd.dgstd)
    return Response({'applied_dgstd':response.data})

@api_view(['POST'])
def postDiagStd(request, format=None):
    print(request)
    data = request.data['form']
    data = getPinYin(data)
    response = None
    print('data:', data)
    dgstd_serializer = DiagStdSerializer(data=data)
    print('validation:', dgstd_serializer.is_valid())
    if dgstd_serializer.is_valid():
        response = DiagStdSerializer(dgstd_serializer.save())
    else:
        print(dgstd_serializer.errors)
    print('giao')
    print(response)
    print(type(response))
    return Response({"dgstd":response.data})

@api_view(['POST'])
def postDiag(request, format=None):
    data = request.data['form']
    # {
    #     code: '',
    #     label: '',
    #     pinyin: '',
    #     pinyin_cap: '',
    #     description: '',
    # }
    print('data:',data)
    diag_serializer = DiagSerializer(data=data)
    print('validation:',diag_serializer.is_valid())
    if diag_serializer.is_valid():
        diag_serializer.save()
    else:
        print(diag_serializer.errors)
    print('giao')
    
    return Response({"request.data":request.data})

@api_view(['POST'])
def updateDiagStd(request):
    data = request.data
    print(data)
    data = getPinYin(data)
    print('id:', data['id']) # id?
    print('name:', data['name'])
    print('dgstd:',DiagStd.objects.all())
    instance = DiagStd.objects.get(id__exact=data['id'])
    print('instance:',instance)
    # instance.delete()
    dgstd_serializer = DiagStdSerializer(instance, data)
    print('new dgstd valid:',dgstd_serializer.is_valid())
    if dgstd_serializer.is_valid():
        dgstd_serializer.save()
    else:
        print(dgstd_serializer.errors)
    print('giao')
    
    return Response({"request.data":request.data})

def getPinYin(data):
    print('@getPinYin()')
    print(data)
    for diag in data['diag']:
        print('diag:',diag)
        print('diag[label]',diag['label'])
        diag['pinyin'] = ''.join(map(lambda p:p[0],\
            pinyin(diag['label'], style=Style.NORMAL)))
        diag['pinyin_cap'] = ''.join(map(lambda p:p[0],\
            pinyin(diag['label'], style=Style.FIRST_LETTER)))
        print(diag['pinyin'])
        print(diag['pinyin_cap'])
    return data 

@api_view(['POST'])
def removeDiagStd(request):
    id = request.data['id']
    print('id', id)
    # print(DiagStd.objects.all())
    instance = DiagStd.objects.get(id__exact=id)
    # instance = DiagStd.objects.get(id=id)
    print('instance',instance)
    if instance:
        print(instance)
        instance.delete()
        print('deleted')
        return Response({"result":str(id)+" removed"})
    else:
        return Response({"result":"something went wrong"})

class ViewDgStd(APIView):
    def get(self, request, format=None):
        id = request.GET.get('id')
        print('spstd w/ id:', id)
        stand = DiagStd.objects.get(id__exact=id)
        print(stand.id)
        serializer = DiagStdSerializer(stand, many=False)

        return Response({'dgstd':serializer.data})

class DgStdList(APIView):
    def get(self, request, format=None):
        print('@getDiagStdList()')
        dgstd_list = DiagStd.objects.all()
        print(dgstd_list)
        serializer = DiagStdSerializer(dgstd_list, many=True)
        print(serializer.data)
        return Response({'dgstd_list':serializer.data})

class DiagQuery(APIView):
    def get(self, request, format=None):
        print('@queryDiag()')
        print('request:',request)
        query = request.GET.get('query')
        print('query:',query)
        query = ''.join(map(lambda c:c[0],pinyin(query, style=Style.NORMAL)))
        print('query:',query)
        cursor = connection.cursor()
        sql = "select * from standard_diag where (pinyin like concat('%s','%%') or pinyin_cap like concat('%s','%%')) and diagstd_id=(select id from standard_diagstd where id=(select dgstd_id from standard_applieddgstd))"%(query,query)
        print(sql)
        cursor.execute(sql)
        records = cursor.fetchall()
        print(type(records))
        print('records:',records)
        result = []
        if len(records)!=0:
            for r in records:
                result.append({
                    'code':r[1],
                    'label':r[2],
                    'pinyin':r[3],
                    'pinyin_cap':r[4],
                    'description':r[5]
                })
        return Response({'result':result})

# post std
# @api_view(['POST'])
# setappliedstd
# getstd
# getstds
# getappliedstd
# updatestd
# removestd
@api_view(['GET'])
def getAppliedGStd(request):
    type = request.GET.get('type')
    typed_gstd = GStd.objects.filter(type=type)[:1]
    if len(typed_gstd)==0: # 如无该类标准
        gstd = GStd.objects.create(name='默认标准',type=type)
    else: # 如有该类标准
        gstd = typed_gstd.get()
    applied_gstd = None
    if not AppliedGStds.objects.exists(): # 如无AppliedGStds
        print('---- creating AppliedGStd of type',type)
        print('gstd got:',gstd)
        # 根据type新建携带改类标准的AppliedGStds
        if type == 'NATIONALITY':
            applied_gstd = AppliedGStds.objects.create(nationality_std=gstd)
        elif type == 'ETHNICITY':
            applied_gstd = AppliedGStds.objects.create(ethnicity_std=gstd)
        elif type == 'IDTYPE':
            applied_gstd = AppliedGStds.objects.create(id_type_std=gstd)
        elif type == 'GENDER':
            applied_gstd = AppliedGStds.objects.create(gender_std=gstd)
        elif type == 'MARRIAGESTAT':
            applied_gstd = AppliedGStds.objects.create(marriage_stat_std=gstd)
        elif type == 'SETTLEMENTTYPE':
            applied_gstd = AppliedGStds.objects.create(settlemenet_type_std=gstd)
        elif type == 'CONTACTRELATION':
            applied_gstd = AppliedGStds.objects.create(contact_relation_std=gstd)
        elif type == 'SPECIALPERSONTYPE':
            applied_gstd = AppliedGStds.objects.create(special_person_type_std=gstd)
        elif type == 'NEWBORNADMITTYPE':
            applied_gstd = AppliedGStds.objects.create(newborn_admit_type_std=gstd)
        elif type == 'HOSPREASON':
            applied_gstd = AppliedGStds.objects.create(hosp_reason_std=gstd)
        elif type == 'HEALTYPE':
            applied_gstd = AppliedGStds.objects.create(heal_type_std=gstd)
        elif type == 'ADMITPATH':
            applied_gstd = AppliedGStds.objects.create(admit_path_std=gstd)
        elif type == 'ANAESTHESIATYPE':
            applied_gstd = AppliedGStds.objects.create(anaesthesia_type_std=gstd)
        elif type == 'CUTYPE':
            applied_gstd = AppliedGStds.objects.create(cu_type_std=gstd)
        elif type == 'BLOODTYPE':
            applied_gstd = AppliedGStds.objects.create(blood_type_std=gstd)
        elif type == 'PAYMENTTYPE':
            applied_gstd = AppliedGStds.objects.create(payment_type_std=gstd)
        elif type == 'PURCHASEMETHOD':
            applied_gstd = AppliedGStds.objects.create(purchase_method_std=gstd)
        elif type == 'ADMITCONDITION':
            applied_gstd = AppliedGStds.objects.create(admit_condition_std=gstd)
        elif type == 'BLOODGROUP':
            applied_gstd = AppliedGStds.objects.create(blood_group_std=gstd)
        elif type == 'RH':
            applied_gstd = AppliedGStds.objects.create(rh_std=gstd)
        elif type == 'RECORDQUALITY':
            applied_gstd = AppliedGStds.objects.create(record_quality_std=gstd)
        elif type == 'OPLVL':
            applied_gstd = AppliedGStds.objects.create(op_lvl_std=gstd)
        elif type == 'WOUNDHEALINGLVL':
            applied_gstd = AppliedGStds.objects.create(wound_healing_lvl_std=gstd)
        elif type == 'RELEASETYPE':
            applied_gstd = AppliedGStds.objects.create(release_type_std=gstd)
        print('test')
    else: # 如有AppliedGStds
        applied_gstd = AppliedGStds.objects.get()
        # 检查有无该类标准并添加
        if type == 'NATIONALITY':
            if not applied_gstd.nationality_std:
                applied_gstd.nationality_std = gstd
        elif type == 'ETHNICITY':
            if not applied_gstd.ethnicity_std:
                applied_gstd.ethnicity_std = gstd
        elif type == 'IDTYPE':
            if not applied_gstd.id_type_std:
                applied_gstd.id_type_std = gstd
        elif type == 'GENDER':
            if not applied_gstd.gender_std:
                applied_gstd.gender_std = gstd
        elif type == 'MARRIAGESTAT':
            if not applied_gstd.marriage_stat_std:
                applied_gstd.marriage_stat_std = gstd
        elif type == 'SETTLEMENTTYPE':
            if not applied_gstd.settlemenet_type_std:
                applied_gstd.settlemenet_type_std = gstd
        elif type == 'CONTACTRELATION':
            if not applied_gstd.contact_relation_std:
                applied_gstd.contact_relation_std = gstd
        elif type == 'SPECIALPERSONTYPE':
            if not applied_gstd.special_person_type_std:
                applied_gstd.special_person_type_std = gstd
        elif type == 'NEWBORNADMITTYPE':
            if not applied_gstd.newborn_admit_type_std:
                applied_gstd.newborn_admit_type_std = gstd
        elif type == 'HOSPREASON':
            if not applied_gstd.hosp_reason_std:
                applied_gstd.hosp_reason_std = gstd
        elif type == 'HEALTYPE':
            if not applied_gstd.heal_type_std:
                applied_gstd.heal_type_std = gstd
        elif type == 'ADMITPATH':
            if not applied_gstd.admit_path_std:
                applied_gstd.admit_path_std = gstd
        elif type == 'ANAESTHESIATYPE':
            if not applied_gstd.anaesthesia_type_std:
                applied_gstd.anaesthesia_type_std = gstd
        elif type == 'CUTYPE':
            if not applied_gstd.cu_type_std:
                applied_gstd.cu_type_std = gstd
        elif type == 'BLOODTYPE':
            if not applied_gstd.blood_type_std:
                applied_gstd.blood_type_std = gstd
        elif type == 'PAYMENTTYPE':
            if not applied_gstd.payment_type_std:
                applied_gstd.payment_type_std = gstd
        elif type == 'PURCHASEMETHOD':
            if not applied_gstd.purchase_method_std:
                applied_gstd.purchase_method_std = gstd
        elif type == 'ADMITCONDITION':
            if not applied_gstd.admit_condition_std:
                applied_gstd.admit_condition_std = gstd
        elif type == 'BLOODGROUP':
            if not applied_gstd.blood_group_std:
                applied_gstd.blood_group_std = gstd
        elif type == 'RH':
            if not applied_gstd.rh_std:
                applied_gstd.rh_std = gstd
        elif type == 'RECORDQUALITY':
            if not applied_gstd.record_quality_std:
                applied_gstd.record_quality_std = gstd
        elif type == 'OPLVL':
            if not applied_gstd.op_lvl_std:
                applied_gstd.op_lvl_std = gstd
        elif type == 'WOUNDHEALINGLVL':
            if not applied_gstd.wound_healing_lvl_std:
                applied_gstd.wound_healing_lvl_std = gstd
        elif type == 'RELEASETYPE':
            if not applied_gstd.release_type_std:
                applied_gstd.release_type_std = gstd
    print('applied_stds got:',applied_gstd)
    # 根据type决定resopnse
    if type == 'NATIONALITY':
        response = GStdSerializer(applied_gstd.nationality_std)
    elif type == 'ETHNICITY':
        response = GStdSerializer(applied_gstd.ethnicity_std)
    elif type == 'IDTYPE':
        response = GStdSerializer(applied_gstd.id_type_std)
    elif type == 'GENDER':
        response = GStdSerializer(applied_gstd.gender_std)
    elif type == 'MARRIAGESTAT':
        response = GStdSerializer(applied_gstd.marriage_stat_std)
    elif type == 'SETTLEMENTTYPE':
        response = GStdSerializer(applied_gstd.settlemenet_type_std)
    elif type == 'CONTACTRELATION':
        response = GStdSerializer(applied_gstd.contact_relation_std)
    elif type == 'SPECIALPERSONTYPE':
        response = GStdSerializer(applied_gstd.special_person_type_std)
    elif type == 'NEWBORNADMITTYPE':
        response = GStdSerializer(applied_gstd.newborn_admit_type_std)
    elif type == 'HOSPREASON':
        response = GStdSerializer(applied_gstd.hosp_reason_std)
    elif type == 'HEALTYPE':
        response = GStdSerializer(applied_gstd.heal_type_std)
    elif type == 'ADMITPATH':
        response = GStdSerializer(applied_gstd.admit_path_std)
    elif type == 'ANAESTHESIATYPE':
        response = GStdSerializer(applied_gstd.anaesthesia_type_std)
    elif type == 'CUTYPE':
        response = GStdSerializer(applied_gstd.cu_type_std)
    elif type == 'BLOODTYPE':
        response = GStdSerializer(applied_gstd.blood_type_std)
    elif type == 'PAYMENTTYPE':
        response = GStdSerializer(applied_gstd.payment_type_std)
    elif type == 'PURCHASEMETHOD':
        response = GStdSerializer(applied_gstd.purchase_method_std)
    elif type == 'ADMITCONDITION':
        response = GStdSerializer(applied_gstd.admit_condition_std)
    elif type == 'BLOODGROUP':
        response = GStdSerializer(applied_gstd.blood_group_std)
    elif type == 'RH':
        response = GStdSerializer(applied_gstd.rh_std)
    elif type == 'RECORDQUALITY':
        response = GStdSerializer(applied_gstd.record_quality_std)
    elif type == 'OPLVL':
        response = GStdSerializer(applied_gstd.op_lvl_std)
    elif type == 'WOUNDHEALINGLVL':
        response = GStdSerializer(applied_gstd.wound_healing_lvl_std)
    elif type == 'RELEASETYPE':
        response = GStdSerializer(applied_gstd.release_type_std)
    print(response.data)
    generals = response.data['general']
    id = response.data['id']
    return Response({'id':id,'generals':generals})

names = dict()
names['NATIONALITY'] = '国籍'
names['ETHNICITY'] = '民族'
names['IDTYPE'] = '证件类型'
names['GENDER'] = '性别'
names['MARRIAGESTAT'] = '婚姻状态'
names['SETTLEMENTTYPE'] = '医保类型'
names['CONTACTRELATION'] = '联系人关系'
names['SPECIALPERSONTYPE'] = '特殊人员类型'
names['NEWBORNADMITTYPE'] = '新生儿入院类型'
names['HOSPREASON'] = '住院医疗类型'
names['HEALTYPE'] = '治疗类别'
names['ADMITPATH'] = '入院途径'
names['ANAESTHESIATYPE'] = '麻醉方式'
names['CUTYPE'] = '重症监护病房类型'
names['BLOODTYPE'] = '输血品种'
names['PAYMENTTYPE'] = '医疗支付方式'
names['PURCHASEMETHOD'] = '医疗付费方式'
names['ADMITCONDITION'] = '入院病情'
names['BLOODGROUP'] = '血型'
names['RH'] = 'Rh'
names['RECORDQUALITY'] = '病案质量'
names['OPLVL'] = '手术级别'
names['WOUNDHEALINGLVL'] = '切口愈合等级'
names['RELEASETYPE'] = '离院方式'

@api_view(['GET'])
def getTypeName(request):
    print('request', request.GET)
    return Response(names[request.GET['type']])

@api_view(['GET'])
def getAppliedGStds(request):
    # dict : type->typed std variable name
    if not AppliedGStds.objects.exists():
        stds = None
    else:
        applied_gstds = AppliedGStds.objects.get()
        stds = dict()
        stds['NATIONALITY'] = applied_gstds.nationality_std.id if applied_gstds.nationality_std else None
        stds['ETHNICITY'] = applied_gstds.ethnicity_std.id if applied_gstds.ethnicity_std else None
        stds['IDTYPE'] = applied_gstds.id_type_std.id if applied_gstds.id_type_std else None
        stds['GENDER'] = applied_gstds.gender_std.id if applied_gstds.gender_std else None
        stds['MARRIAGESTAT'] = applied_gstds.marriage_stat_std.id if applied_gstds.marriage_stat_std else None
        stds['SETTLEMENTTYPE'] = applied_gstds.settlemenet_type_std.id if applied_gstds.settlemenet_type_std else None
        stds['CONTACTRELATION'] = applied_gstds.contact_relation_std.id if applied_gstds.contact_relation_std else None
        stds['SPECIALPERSONTYPE'] = applied_gstds.special_person_type_std.id if applied_gstds.special_person_type_std else None
        stds['NEWBORNADMITTYPE'] = applied_gstds.newborn_admit_type_std.id if applied_gstds.newborn_admit_type_std else None
        stds['HOSPREASON'] = applied_gstds.hosp_reason_std.id if applied_gstds.hosp_reason_std else None
        stds['HEALTYPE'] = applied_gstds.heal_type_std.id if applied_gstds.heal_type_std else None
        stds['ADMITPATH'] = applied_gstds.admit_path_std.id if applied_gstds.admit_path_std else None
        stds['ANAESTHESIATYPE'] = applied_gstds.anaesthesia_type_std.id if applied_gstds.anaesthesia_type_std else None
        stds['CUTYPE'] = applied_gstds.cu_type_std.id if applied_gstds.cu_type_std else None
        stds['BLOODTYPE'] = applied_gstds.blood_type_std.id if applied_gstds.blood_type_std else None
        stds['PAYMENTTYPE'] = applied_gstds.payment_type_std.id if applied_gstds.payment_type_std else None
        stds['PURCHASEMETHOD'] = applied_gstds.purchase_method_std.id if applied_gstds.purchase_method_std else None
        stds['ADMITCONDITION'] = applied_gstds.admit_condition_std.id if applied_gstds.admit_condition_std else None
        stds['BLOODGROUP'] = applied_gstds.blood_group_std.id if applied_gstds.blood_group_std else None
        stds['RH'] = applied_gstds.rh_std.id if applied_gstds.rh_std else None
        stds['RECORDQUALITY'] = applied_gstds.record_quality_std.id if applied_gstds.record_quality_std else None
        stds['OPLVL'] = applied_gstds.op_lvl_std.id if applied_gstds.op_lvl_std else None
        stds['WOUNDHEALINGLVL'] = applied_gstds.wound_healing_lvl_std.id if applied_gstds.wound_healing_lvl_std else None
        stds['RELEASETYPE'] = applied_gstds.release_type_std.id if applied_gstds.release_type_std else None
    # 添加type->中文的对照，用于前端显示

    response = dict()
    response['standards'] = stds
    response['names'] = names
    print(response)
    return Response(response)

@api_view(['POST'])
def setAppliedGStd(request, format=None):
    print('setAppliedGStd()')
    print('request.data',request.data)
    id = request.data['id']
    type = request.data['type']
    gstd = GStd.objects.get(id__exact=id)
    
    if not AppliedGStds.objects.exists(): # 如无AppliedGStds
        print('---- creating AppliedStd')
        applied_gstd = AppliedGStds.objects.create()
    else: # 如有AppliedGStds
        applied_gstd = AppliedGStds.objects.get()
    # 根据type设置gstd
    if type == 'NATIONALITY':
        applied_gstd.nationality_std = gstd
    elif type == 'ETHNICITY':
        applied_gstd.ethnicity_std = gstd
    elif type == 'IDTYPE':
        applied_gstd.id_type_std = gstd
    elif type == 'GENDER':
        applied_gstd.gender_std = gstd
    elif type == 'MARRIAGESTAT':
        applied_gstd.marriage_stat_std = gstd
    elif type == 'SETTLEMENTTYPE':
        applied_gstd.settlemenet_type_std = gstd
    elif type == 'CONTACTRELATION':
        applied_gstd.contact_relation_std = gstd
    elif type == 'SPECIALPERSONTYPE':
        applied_gstd.special_person_type_std = gstd
    elif type == 'NEWBORNADMITTYPE':
        applied_gstd.newborn_admit_type_std = gstd
    elif type == 'HOSPREASON':
        applied_gstd.hosp_reason_std = gstd
    elif type == 'HEALTYPE':
        applied_gstd.heal_type_std = gstd
    elif type == 'ADMITPATH':
        applied_gstd.admit_path_std = gstd
    elif type == 'ANAESTHESIATYPE':
        applied_gstd.anaesthesia_type_std = gstd
    elif type == 'CUTYPE':
        applied_gstd.cu_type_std = gstd
    elif type == 'BLOODTYPE':
        applied_gstd.blood_type_std = gstd
    elif type == 'PAYMENTTYPE':
        applied_gstd.payment_type_std = gstd
    elif type == 'PURCHASEMETHOD':
        applied_gstd.purchase_method_std = gstd
    elif type == 'ADMITCONDITION':
        applied_gstd.admit_condition_std = gstd
    elif type == 'BLOODGROUP':
        applied_gstd.blood_group_std = gstd
    elif type == 'RH':
        applied_gstd.rh_std = gstd
    elif type == 'RECORDQUALITY':
        applied_gstd.record_quality_std = gstd
    elif type == 'OPLVL':
        applied_gstd.op_lvl_std = gstd
    elif type == 'WOUNDHEALINGLVL':
        applied_gstd.wound_healing_lvl_std = gstd
    elif type == 'RELEASETYPE':
        applied_gstd.release_type_std = gstd
    applied_gstd.save()
    print('applied_dgstd got:',applied_gstd)
    
    response = GStdSerializer(gstd)
    return Response({'applied_gstd':response.data})

@api_view(['POST'])
def postGStd(request, format=None):
    print(request)
    data = request.data['form']
    response = None
    print('data:', data)
    gstd_serializer = GStdSerializer(data=data)
    print('validation:', gstd_serializer.is_valid())
    if gstd_serializer.is_valid():
        response = GStdSerializer(gstd_serializer.save())
    else:
        print(gstd_serializer.errors)
    print('giao')
    print(response)
    print(type(response))
    return Response({"gstd":response.data})

@api_view(['POST'])
def updateGStd(request):
    print('@updateGStd()')
    print('request:',request)
    data = request.data
    id = data['id']
    type = data['type']
    print(data)
    print('id:', id) # id?
    print('type:', type)
    instance = GStd.objects.get(id__exact=id)
    print('instance:',instance)
    gstd_serializer = GStdSerializer(instance, data)
    print('new gstd valid:',gstd_serializer.is_valid())
    if gstd_serializer.is_valid():
        gstd_serializer.save()
    else:
        print(gstd_serializer.errors)
    print('giao')
    return Response({"request.data":request.data})

@api_view(['POST'])
def postGeneral(request, format=None):
    data = request.data['form']
    print('data:',data)
    general_serializer = GeneralSerializer(data=data)
    print('validation:',general_serializer.is_valid())
    if general_serializer.is_valid():
        general_serializer.save()
    else:
        print(general_serializer.errors)
    print('giao')
    
    return Response({"request.data":request.data})

class ViewGStd(APIView):
    def get(self, request, format=None):
        id = request.GET.get('id')
        print('spstd w/ id:', id)
        stand = GStd.objects.get(id__exact=id)
        print(stand.id)
        serializer = GStdSerializer(stand, many=False)
        response = dict()
        response['gstd'] = serializer.data
        response['type_name'] = names[stand.type]
        return Response(response)

class GStdList(APIView):
    def get(self, request, format=None):
        print('@getGStdList()')
        type = request.GET.get('type')
        print('type:', type)
        # gstd_list = GStd.objects.all()
        gstd_list = GStd.objects.filter(type=type)
        print(gstd_list)
        serializer = GStdSerializer(gstd_list, many=True)
        print(serializer.data)
        return Response({'gstd_list':serializer.data})

class AllGStdList(APIView):
    def get(self, request, format=None):
        print('@getNoTypeGStdList()')
        gstd_list = GStd.objects.all()
        print(gstd_list)
        serializer = GStdSerializer(gstd_list, many=True)
        print(serializer.data)
        return Response({'gstd_list':serializer.data})

@api_view(['POST'])
def removeGStd(request):
    id = request.data['id']
    print('id', id)
    instance = GStd.objects.get(id__exact=id)
    print('instance',instance)
    if instance:
        print(instance)
        instance.delete()
        print('deleted')
        return Response({"result":str(id)+" removed"})
    else:
        return Response({"result":"something went wrong"})

class UploadGXlsViewSet(ModelViewSet):
    queryset = UploadModel.objects.all()
    def create(self, request):
        # 检查类型信息在哪里放置
        type = request.data['type']
        print('request',request)
        # type = request
        file_dir = None
        print('UploadGXlsViewSet.create()')
        upload_serializer = UploadSerializer(data=request.data)
        print('validation:', upload_serializer.is_valid())
        if(upload_serializer.is_valid()):
            upload = upload_serializer.save()
        file_name = str(upload.file).split('/')[1]
        print('file_name:', file_name)
        file_dir = str(settings.MEDIA_ROOT)+'\\uploads\\'+file_name
        print('file_dir:', file_dir)
        output = dict(name=file_name, general=[],type=type)
        file_type = file_name.split('.')[1]
        print('file_type:',file_type)
        if(file_type=='xls'):
            file = xlrd.open_workbook(file_dir)
            sheet = file.sheets()[0]
            print('sheet:',sheet)
            print(sheet.nrows)
            for i in range(1,sheet.nrows):
                code = str(sheet.cell_value(i,0)).strip(' ')
                label = sheet.cell_value(i,1).strip(' ')
                output['general'].append(dict(code=code,label=label))
        elif(file_type=='xlsx'):
            file = openpyxl.load_workbook(file_dir)
            sheets = file.sheetnames
            sheet = file[sheets[0]]
            print('max_row:',sheet.max_row)
            for i in range(2,sheet.max_row + 1):
                if sheet.cell(column=1, row=i).value==None or sheet.cell(column=2, row=i).value==None:
                    continue
                code = str(sheet.cell(column=1, row=i).value).strip(' ')
                label = sheet.cell(column=2, row=i).value.strip(' ')
                print(code+' -- '+label+'     >> '+str(len(code)))
                temp = dict(code=code,label=label)
                output['general'].append(temp)
        else:
            return Response("有问题啊")
        gstd_serializer = GStdSerializer(data=output)
        print('validation:', gstd_serializer.is_valid())
        if gstd_serializer.is_valid():
            gstd_serializer.save()
        else:
            print(gstd_serializer.errors)
        print('giao')
        response = "POST API: 你一上传了一个文件:{}".format(file_name)
        return Response(response)

class UploadGJsonViewSet(ModelViewSet):
    queryset = UploadModel.objects.all()
    def create(self, request):
        print(request.data)
        type = request.data['type']
        file_dir = None # 文件存储路径
        json_str = None # 读取的json
        print('UploadGJsonViewSet.create()')
        upload_serializer = UploadSerializer(data=request.data)
        print('validation:', upload_serializer.is_valid())
        if(upload_serializer.is_valid()):
            upload = upload_serializer.save()
        file_name = str(upload.file).split('/')[1]
        file_dir = str(settings.MEDIA_ROOT)+'\\uploads\\'+file_name
        print('file_dir:', file_dir)
        # 打开json文件
        json_str = None
        with open(file_dir, encoding = 'utf-8') as f:
            json_str = json.load(f)
        if json_str:
            json_str['general'] = json_str[type.lower()]
            json_str['name'] = file_name
            json_str['type'] = type
            gstd_serializer = GStdSerializer(data=json_str)
            print('validation:', gstd_serializer.is_valid())
            if gstd_serializer.is_valid():
                gstd_serializer.save()
            else:
                print(gstd_serializer.errors)
            print('giao')
        response = "POST API: 你一上传了一个文件:{}".format(file_name)
        return Response(response)